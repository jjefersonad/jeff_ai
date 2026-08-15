"""Testes do XlsxWriter (infrastructure/xlsx_writer) e da tool create_xlsx_spreadsheet.

Cobrem os critérios de aceitação da task `custom-office-doc-tools-task-xlsx-1`:
- REQ-001: gera `.xlsx` válido só com openpyxl, sem binários externos.
- REQ-002: suporta múltiplas abas, células, cabeçalhos, fórmulas e formatação
  básica (negrito, largura de coluna, formato numérico).
- REQ-003: salva em `outputs/documents/xlsx/<ts>.xlsx`, retorna
  `{path, url, metadata}` e a URL começa com `/api/files/xlsx/`.
- REQ-005: entrada inválida retorna erro descritivo e não deixa arquivo parcial.
- Tool como adapter fino + wiring em `composition.dependencies`.
"""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

import src.composition.dependencies as dep
import src.tools.create_xlsx_spreadsheet_tool as xlsx_tool
from src.domain.documents import Sheet, XlsxSpec
from src.domain.shared.errors import DomainError
from src.infrastructure.documents.xlsx_writer import XlsxWriter
from src.models.xlsx_document import XlsxDocumentInput, XlsxSheetInput

# --- XlsxWriter (infra) ----------------------------------------------------


@pytest.fixture
def writer(tmp_path: Path) -> XlsxWriter:
    """XlsxWriter apontando para um diretório temporário (sem poluir outputs/)."""
    return XlsxWriter(output_dir=tmp_path, url_prefix="/api/files")


async def test_writer_creates_valid_xlsx_with_two_sheets_and_formula(writer, tmp_path):
    """REQ-001/REQ-002/REQ-003: writer gera .xlsx válido, 2 abas, fórmula preservada."""
    spec = XlsxSpec(
        sheets=(
            Sheet(
                name="Vendas",
                rows=(
                    ("Mês", "Receita", "Custo"),
                    ("Jan", 12000, 8000),
                    ("Fev", 15000, 9000),
                    ("Total", "=SUM(B2:B3)", "=SUM(C2:C3)"),
                ),
                header=True,
            ),
            Sheet(name="Resumo", rows=(("OK",),)),
        )
    )

    result = await writer.write(spec)

    # REQ-003: contrato {path, url, metadata}, URL pública e arquivo no disco.
    assert result.path
    assert result.url.startswith("/api/files/xlsx/")
    assert result.url.endswith(".xlsx")
    assert result.metadata == {"kind": "xlsx", "sheets": ["Vendas", "Resumo"]}
    path = Path(result.path)
    assert path.is_file()
    assert path.parent == tmp_path

    # Reabre com openpyxl e valida estrutura.
    workbook = load_workbook(str(path))
    assert workbook.sheetnames == ["Vendas", "Resumo"]

    vendas = workbook["Vendas"]
    assert vendas["A1"].value == "Mês"
    assert vendas["B2"].value == 12000
    assert vendas["A4"].value == "Total"
    # Fórmula preservada como string iniciada por '='.
    assert vendas["B4"].value == "=SUM(B2:B3)"
    assert vendas["C4"].value == "=SUM(C2:C3)"

    resumo = workbook["Resumo"]
    assert resumo["A1"].value == "OK"


async def test_writer_applies_header_bold_and_column_widths(writer, tmp_path):
    """REQ-002: cabeçalho em negrito + largura de colunas."""
    spec = XlsxSpec(
        sheets=(
            Sheet(
                name="Dados",
                rows=(("Col 1", "Col 2"), ("v1", "v2")),
                header=True,
                column_widths=(20.0, 30.0),
            ),
        )
    )

    result = await writer.write(spec)
    workbook = load_workbook(result.path)
    sheet = workbook["Dados"]

    # Cabeçalho em negrito: células da primeira linha.
    assert sheet["A1"].font.bold is True
    assert sheet["B1"].font.bold is True
    # Linhas de dados: não-negrito.
    assert sheet["A2"].font.bold is None or sheet["A2"].font.bold is False

    # Larguras de coluna aplicadas.
    assert sheet.column_dimensions["A"].width == 20.0
    assert sheet.column_dimensions["B"].width == 30.0


async def test_writer_applies_number_format(writer, tmp_path):
    """REQ-002: formato numérico por coluna."""
    spec = XlsxSpec(
        sheets=(
            Sheet(
                name="Valores",
                rows=((123.456,), (789.012,)),
                number_formats=((0, "#,##0.00"),),
            ),
        )
    )

    result = await writer.write(spec)
    workbook = load_workbook(result.path)
    sheet = workbook["Valores"]

    assert sheet["A1"].number_format == "#,##0.00"
    assert sheet["A2"].number_format == "#,##0.00"


async def test_writer_keeps_three_sheets_in_order(writer, tmp_path):
    """REQ-002: três abas, ordem preservada."""
    spec = XlsxSpec(
        sheets=(
            Sheet(name="Alpha", rows=(("a",),)),
            Sheet(name="Beta", rows=(("b",),)),
            Sheet(name="Gamma", rows=(("g",),)),
        )
    )

    result = await writer.write(spec)
    workbook = load_workbook(result.path)
    assert workbook.sheetnames == ["Alpha", "Beta", "Gamma"]


async def test_writer_rejects_non_xlsx_spec():
    """O writer só aceita XlsxSpec — typecheck em tempo de execução."""
    writer = XlsxWriter()
    with pytest.raises(TypeError, match="XlsxSpec"):
        await writer.write("not a spec")  # type: ignore[arg-type]


# --- Domínio: validações que protegem o writer ----------------------------


def test_xlsx_spec_rejects_empty_sheets():
    with pytest.raises(DomainError, match="ao menos uma aba"):
        XlsxSpec(sheets=())


def test_sheet_rejects_invalid_cell_type():
    with pytest.raises(DomainError, match="str, int, float ou None"):
        Sheet(name="S", rows=(("ok", [1, 2]),))


def test_sheet_rejects_non_positive_column_width():
    with pytest.raises(DomainError, match="positivos"):
        Sheet(name="S", rows=(("a",),), column_widths=(-1.0, 5.0))


def test_sheet_rejects_negative_column_index_in_number_format():
    with pytest.raises(DomainError, match=">= 0"):
        Sheet(name="S", rows=(("a",),), number_formats=((-1, "0.00"),))


def test_sheet_rejects_empty_name():
    with pytest.raises(DomainError, match="name"):
        Sheet(name="   ", rows=(("a",),))


def test_sheet_rejects_empty_rows():
    """REQ-006 (xlsx-generation): aba presente mas sem linhas é barrada no domínio."""
    with pytest.raises(DomainError, match="rows"):
        Sheet(name="Vendas", rows=())


def test_sheet_accepts_non_empty_rows():
    """REQ-006 (regressão): aba com ao menos uma linha continua funcionando."""
    sheet = Sheet(name="Vendas", rows=(("Mês", "Receita"), ("Jan", 12000)), header=True)
    assert sheet.rows == (("Mês", "Receita"), ("Jan", 12000))


# --- Tool create_xlsx_spreadsheet (HTML pipeline) --------------------------


def test_sheets_to_html_basic():
    html = xlsx_tool.sheets_to_html(
        [
            XlsxSheetInput(name="A", rows=[["x", 1, 1.5, None]]),
            XlsxSheetInput(name="B", rows=[["k"]], header=True),
        ]
    )
    assert 'data-sheet-name="A"' in html
    assert 'data-sheet-name="B"' in html
    assert "<td>x</td>" in html
    assert "<td>k</td>" in html


def test_sheets_to_html_preserves_formula_strings():
    html = xlsx_tool.sheets_to_html(
        [XlsxSheetInput(name="F", rows=[["=A1+B1", 10]])]
    )
    assert "=A1+B1" in html


async def test_tool_returns_path_url_metadata(monkeypatch, tmp_path):
    """A tool usa o pipeline HTML e devolve o contrato esperado."""
    from unittest.mock import AsyncMock

    monkeypatch.setattr(xlsx_tool, "require_user_docs_dir", AsyncMock(return_value=tmp_path))
    monkeypatch.setattr(xlsx_tool, "_document_url_prefix", lambda: "/api/files")
    monkeypatch.setattr(xlsx_tool, "record_ownership", AsyncMock())

    result = await xlsx_tool.create_xlsx_spreadsheet.coroutine(
        XlsxDocumentInput(
            sheets=[
                XlsxSheetInput(
                    name="S1",
                    rows=[["a", 1], ["b", 2]],
                    header=True,
                ),
            ],
        )
    )

    assert set(result) == {"path", "url", "metadata"}
    assert "/api/files/xlsx/" in result["url"]
    assert result["url"].endswith(".xlsx")
    assert result["metadata"]["kind"] == "xlsx"
    assert Path(result["path"]).is_file()


async def test_tool_returns_error_on_invalid_input(monkeypatch, tmp_path):
    """REQ-005: entrada inválida vira `error` descritivo, sem arquivo parcial."""
    from unittest.mock import AsyncMock

    monkeypatch.setattr(xlsx_tool, "require_user_docs_dir", AsyncMock(return_value=tmp_path))
    monkeypatch.setattr(xlsx_tool, "_document_url_prefix", lambda: "/api/files")
    monkeypatch.setattr(xlsx_tool, "record_ownership", AsyncMock())

    result = await xlsx_tool.create_xlsx_spreadsheet.coroutine(
        XlsxDocumentInput(sheets=[]),
    )

    assert "error" in result
    assert "sheets" in result["error"].lower()
    assert list(tmp_path.iterdir()) == []


async def test_tool_uses_html_xlsx_converter(monkeypatch, tmp_path):
    """A tool aponta o pipeline para HtmlXlsxConverter (não XlsxWriter direto)."""
    from unittest.mock import AsyncMock

    from src.infrastructure.documents.html_xlsx_converter import HtmlXlsxConverter

    monkeypatch.setattr(xlsx_tool, "require_user_docs_dir", AsyncMock(return_value=tmp_path))
    monkeypatch.setattr(xlsx_tool, "_document_url_prefix", lambda: "/api/files")
    monkeypatch.setattr(xlsx_tool, "record_ownership", AsyncMock())

    render = xlsx_tool._build_xlsx_render(tmp_path)
    assert isinstance(render._converters["xlsx"], HtmlXlsxConverter)

    result = await xlsx_tool.create_xlsx_spreadsheet.coroutine(
        XlsxDocumentInput(sheets=[XlsxSheetInput(name="S", rows=[["x"]])]),
    )
    assert "error" not in result
    assert Path(result["path"]).is_file()
