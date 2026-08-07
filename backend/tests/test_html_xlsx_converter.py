"""Testes do converter HTML→XLSX (html-document-tools-task-xlsx-1)."""
from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook

from src.domain.shared.errors import DomainError
from src.infrastructure.documents.html_xlsx_converter import HtmlXlsxConverter


@pytest.mark.asyncio
async def test_html_table_becomes_xlsx_cells() -> None:
    html = (
        "<table data-sheet-name='Vendas'>"
        "<tr><th>A</th><th>B</th></tr>"
        "<tr><td>1</td><td>2</td></tr>"
        "</table>"
    )
    converter = HtmlXlsxConverter()
    payload = await converter.convert(html=html, css=None, kind="xlsx")

    assert payload[:2] == b"PK"
    wb = load_workbook(BytesIO(payload))
    assert "Vendas" in wb.sheetnames
    ws = wb["Vendas"]
    assert ws.cell(1, 1).value == "A"
    assert ws.cell(2, 2).value in (2, "2")


@pytest.mark.asyncio
async def test_html_without_table_raises() -> None:
    converter = HtmlXlsxConverter()
    with pytest.raises(DomainError, match="tabela"):
        await converter.convert(html="<p>sem tabela</p>", css=None, kind="xlsx")
