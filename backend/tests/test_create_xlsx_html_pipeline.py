"""create_xlsx_spreadsheet via HTML pipeline (html-document-tools-task-xlsx-1)."""
from __future__ import annotations

from io import BytesIO
from pathlib import Path
from unittest.mock import AsyncMock

import pytest
from openpyxl import load_workbook

import src.tools.create_xlsx_spreadsheet_tool as xlsx_tool
from src.models.html_document_input import HtmlDocumentInput
from src.models.xlsx_document import XlsxDocumentInput, XlsxSheetInput


@pytest.fixture
def documents_root(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Path:
    root = tmp_path / "documents"
    monkeypatch.setattr(xlsx_tool, "_documents_base_dir", lambda: root)
    monkeypatch.setattr(
        xlsx_tool,
        "_document_url_prefix",
        lambda: "http://localhost:3000/api/files",
    )
    monkeypatch.setattr(xlsx_tool, "record_ownership", AsyncMock())
    return root


async def test_html_table_returns_xlsx_url(documents_root: Path) -> None:
    """Unit-1: HTML com tabela → url /api/files/xlsx/ + kind=xlsx + células."""
    out = await xlsx_tool.create_xlsx_spreadsheet.coroutine(
        HtmlDocumentInput(
            html=(
                "<table>"
                "<tr><th>Mês</th><th>Receita</th></tr>"
                "<tr><td>Jan</td><td>12000</td></tr>"
                "</table>"
            ),
            title="Vendas",
        )
    )

    assert "error" not in out
    assert out["metadata"]["kind"] == "xlsx"
    assert "/api/files/xlsx/" in out["url"]
    path = Path(out["path"])
    assert path.is_file()
    assert path.read_bytes()[:2] == b"PK"

    wb = load_workbook(BytesIO(path.read_bytes()))
    ws = wb.active
    assert ws.cell(1, 1).value == "Mês"
    assert ws.cell(1, 2).value == "Receita"
    assert ws.cell(2, 1).value == "Jan"
    assert ws.cell(2, 2).value in (12000, "12000")


async def test_no_table_rejects_without_file(documents_root: Path) -> None:
    """Unit-2: HTML sem tabela → error, sem xlsx."""
    out = await xlsx_tool.create_xlsx_spreadsheet.coroutine(
        HtmlDocumentInput(html="<p>Só texto, sem tabela.</p>", title="Vazio")
    )

    assert "error" in out
    assert "path" not in out
    xlsx_dir = documents_root / "xlsx"
    assert not xlsx_dir.exists() or list(xlsx_dir.glob("*.xlsx")) == []


async def test_legacy_sheets_still_work(documents_root: Path) -> None:
    out = await xlsx_tool.create_xlsx_spreadsheet.coroutine(
        XlsxDocumentInput(
            sheets=[
                XlsxSheetInput(
                    name="Vendas",
                    rows=[["Mês", "Receita"], ["Jan", 12000]],
                    header=True,
                )
            ]
        )
    )
    assert "error" not in out
    assert out["metadata"]["kind"] == "xlsx"
    wb = load_workbook(out["path"])
    assert wb.active.cell(1, 1).value == "Mês"
