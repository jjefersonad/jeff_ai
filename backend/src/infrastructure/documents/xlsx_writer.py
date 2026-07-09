"""Adapter de criação de planilhas `.xlsx` via openpyxl (implementa o port)."""
from __future__ import annotations

import asyncio
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from src.application.ports.document_writer import DocumentWriterPort
from src.domain.documents import DocumentResult, DocumentSpec, Sheet, XlsxSpec
from src.infrastructure.documents.output_target import DocumentOutput


class XlsxWriter(DocumentWriterPort):
    """Gera um `.xlsx` a partir de um `XlsxSpec` usando apenas openpyxl.

    Não invoca soffice/pandoc/node. Persiste em `outputs/documents/xlsx/` e monta
    a URL pública (`/api/files/xlsx/...`). O workbook é montado em memória e só é
    escrito no final (`save`), então falhas durante a montagem não deixam arquivo
    parcial.
    """

    def __init__(
        self,
        *,
        output_dir: Path | None = None,
        url_prefix: str = "/api/files",
    ) -> None:
        """Configura o destino e o prefixo de URL do writer."""
        self._output = DocumentOutput("xlsx", output_dir=output_dir, url_prefix=url_prefix)

    async def write(self, spec: DocumentSpec) -> DocumentResult:
        """Gera o `.xlsx` do `spec` (fora do event loop) e retorna o resultado."""
        if not isinstance(spec, XlsxSpec):
            raise TypeError("XlsxWriter só aceita XlsxSpec.")
        return await asyncio.to_thread(self._write_sync, spec)

    def _write_sync(self, spec: XlsxSpec) -> DocumentResult:
        workbook = Workbook()
        # Remove a aba default criada pelo openpyxl — a primeira aba do spec assume.
        default = workbook.active
        workbook.remove(default)

        for sheet in spec.sheets:
            self._render_sheet(workbook, sheet)

        path, url = self._output.allocate(spec.extension)
        workbook.save(str(path))
        return DocumentResult(path=str(path), url=url, metadata=spec.metadata())

    @staticmethod
    def _render_sheet(workbook: Workbook, sheet: Sheet) -> None:
        """Cria uma aba no workbook e popula com linhas/formatação do `Sheet`."""
        worksheet = workbook.create_sheet(title=sheet.name)

        for row in sheet.rows:
            worksheet.append(list(row))

        if sheet.header and worksheet.max_row >= 1:
            header_font = Font(bold=True)
            for cell in worksheet[1]:
                cell.font = header_font

        if sheet.column_widths:
            for col_idx, width in enumerate(sheet.column_widths, start=1):
                worksheet.column_dimensions[_column_letter(col_idx)].width = width

        for col_idx, number_format in sheet.number_formats:
            for row in worksheet.iter_rows(min_col=col_idx + 1, max_col=col_idx + 1):
                for cell in row:
                    cell.number_format = number_format


def _column_letter(col_idx: int) -> str:
    """Convert 1-based index to Excel-style column letter (A, B, ..., AA)."""
    letters: list[str] = []
    n = col_idx
    while n > 0:
        n, rem = divmod(n - 1, 26)
        letters.append(chr(ord("A") + rem))
    return "".join(reversed(letters))


# Re-export alinhado para evitar import não-utilizado em alguns linters.
__all__ = ["XlsxWriter"]
