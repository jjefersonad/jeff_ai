"""Converter semântico HTML `<table>` → XLSX via openpyxl (stdlib HTMLParser).

CSS é ignorado nesta v1 — conversão tabular, não fidelidade de estilo.
"""
from __future__ import annotations

from html.parser import HTMLParser
from io import BytesIO

from openpyxl import Workbook

from src.application.ports.html_document_converter import HtmlDocumentConverter
from src.domain.shared.errors import DomainError

_SKIP_TAGS = frozenset({"script", "style", "head", "title"})
_VOID_SKIP_TAGS = frozenset({"meta", "link"})


def _coerce_cell(raw: str) -> str | int | float:
    text = raw.strip()
    if not text:
        return ""
    try:
        if "." in text:
            return float(text)
        return int(text)
    except ValueError:
        return text


class _HtmlTableExtractor(HTMLParser):
    """Extrai tabelas HTML → lista de (sheet_name, rows)."""

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.tables: list[tuple[str, list[list[str | int | float]]]] = []
        self._skip_depth = 0
        self._in_table = False
        self._table_name: str | None = None
        self._rows: list[list[str | int | float]] = []
        self._current_row: list[str | int | float] | None = None
        self._cell_parts: list[str] = []
        self._in_cell = False
        self._table_index = 0

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        lower = tag.lower()
        if lower in _VOID_SKIP_TAGS:
            return
        if lower in _SKIP_TAGS:
            self._skip_depth += 1
            return
        if self._skip_depth:
            return

        attr_map = {k.lower(): (v or "") for k, v in attrs}
        if lower == "table":
            self._in_table = True
            self._table_index += 1
            name = (
                attr_map.get("data-sheet-name")
                or attr_map.get("data-name")
                or f"Sheet{self._table_index}"
            )
            self._table_name = name.strip() or f"Sheet{self._table_index}"
            self._rows = []
            return
        if lower == "tr" and self._in_table:
            self._current_row = []
            return
        if lower in {"td", "th"} and self._in_table:
            self._in_cell = True
            self._cell_parts = []

    def handle_endtag(self, tag: str) -> None:
        lower = tag.lower()
        if lower in _SKIP_TAGS:
            if self._skip_depth:
                self._skip_depth -= 1
            return
        if self._skip_depth:
            return

        if lower in {"td", "th"} and self._in_table and self._in_cell:
            value = _coerce_cell("".join(self._cell_parts))
            if self._current_row is not None:
                self._current_row.append(value)
            self._in_cell = False
            self._cell_parts = []
            return
        if lower == "tr" and self._in_table and self._current_row is not None:
            self._rows.append(self._current_row)
            self._current_row = None
            return
        if lower == "table" and self._in_table:
            name = self._table_name or f"Sheet{self._table_index}"
            if self._rows:
                self.tables.append((name, self._rows))
            self._in_table = False
            self._table_name = None
            self._rows = []

    def handle_data(self, data: str) -> None:
        if self._skip_depth or not self._in_cell:
            return
        self._cell_parts.append(data)


class HtmlXlsxConverter(HtmlDocumentConverter):
    """HTML com `<table>` → bytes XLSX (openpyxl)."""

    async def convert(self, *, html: str, css: str | None, kind: str) -> bytes:
        _ = css
        if kind != "xlsx":
            raise DomainError(
                f"HtmlXlsxConverter só suporta kind='xlsx', recebeu {kind!r}."
            )
        if not html or not html.strip():
            raise DomainError("HTML vazio: nada para converter para XLSX.")

        extractor = _HtmlTableExtractor()
        extractor.feed(html)
        extractor.close()

        if not extractor.tables:
            raise DomainError(
                "HTML sem tabela utilizável para converter para XLSX."
            )

        workbook = Workbook()
        # Remove a aba default; recria a partir das tabelas.
        default = workbook.active
        workbook.remove(default)

        used_names: set[str] = set()
        for name, rows in extractor.tables:
            sheet_name = _unique_sheet_name(name, used_names)
            used_names.add(sheet_name.lower())
            ws = workbook.create_sheet(title=sheet_name)
            for r_idx, row in enumerate(rows, start=1):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()


def _unique_sheet_name(name: str, used: set[str]) -> str:
    """Excel limita nomes a 31 chars e exige unicidade (case-insensitive)."""
    base = (name or "Sheet")[:31]
    candidate = base
    n = 2
    while candidate.lower() in used:
        suffix = f"_{n}"
        candidate = (base[: 31 - len(suffix)] + suffix)[:31]
        n += 1
    return candidate
